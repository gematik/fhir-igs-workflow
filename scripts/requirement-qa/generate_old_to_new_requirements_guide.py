#!/usr/bin/env python3
"""Generate a markdown guide from old requirement IDs to new IG requirements.

Data sources:
- Mapping JSON (default: qa/requirement-mapping.json)
- Excel files (default: .temp/*.xlsx)

Output:
- Markdown page (default: igs/core/input/pagecontent/requirement-mapping-old-to-new.md)

Links:
- Old requirement: https://gemspec.gematik.de/search/index.html?<req-id>
- New requirement: placeholder URL based on mapped file location
  (default placeholder base: {{NEW_IG_BASE_URL}}) //TODO: set base url for new IGs!
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Set
from zipfile import BadZipFile

import openpyxl

OLD_REQ_RE = re.compile(r"^A_\d+(?:-\d+)?$")
OLD_REQ_TOKEN_RE = re.compile(r"A_\d+(?:-\d+)?")
SOURCE_VERSION_SUFFIX_RE = re.compile(r"\s+\d+(?:\.\d+)+$")


def first_non_empty(row: tuple, indexes: List[int]) -> str:
    for idx in indexes:
        if idx >= len(row):
            continue
        value = row[idx]
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def format_source_from_filename(path: Path) -> str:
    stem = path.stem
    if "_V" in stem:
        base, version = stem.rsplit("_V", 1)
        return f"{base} {version}"
    return stem


def normalize_source_name(source: str) -> str:
    text = " ".join(source.split())
    if not text:
        return text
    return SOURCE_VERSION_SUFFIX_RE.sub("", text)


def base_requirement_id(req_id: str) -> str:
    # A_12345-01 -> A_12345 ; A_12345 -> A_12345
    if "-" in req_id:
        return req_id.split("-", 1)[0]
    return req_id


def collect_old_req_metadata(xlsx_paths: Iterable[Path]) -> Dict[str, Dict[str, Set[str]]]:
    metadata: Dict[str, Dict[str, Set[str]]] = defaultdict(
        lambda: {"titles": set(), "sources": set()}
    )

    for path in xlsx_paths:
        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except (BadZipFile, OSError, ValueError):
            # Ignore invalid spreadsheet files in .temp.
            continue
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            header = next(rows, [])
            if not header:
                continue

            index = {
                str(name).strip(): i
                for i, name in enumerate(header)
                if name is not None and str(name).strip()
            }
            if "ID" not in index:
                continue

            id_idx = index["ID"]
            title_indexes = [index[name] for name in ("Titel", "Title") if name in index]
            source_idx = index.get("Quelle (Referenz)")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if id_idx >= len(row):
                    continue
                req_value = row[id_idx]
                if req_value is None:
                    continue

                req_text = str(req_value).strip()
                req_ids = OLD_REQ_TOKEN_RE.findall(req_text)
                if not req_ids:
                    continue

                title = first_non_empty(row, title_indexes)
                source = ""
                if source_idx is not None and source_idx < len(row):
                    source_val = row[source_idx]
                    source = str(source_val).strip() if source_val is not None else ""
                if not source:
                    source = format_source_from_filename(path)
                source = normalize_source_name(source)

                for req_id in req_ids:
                    metadata[req_id]["sources"].add(source)
                    if title:
                        metadata[req_id]["titles"].add(title)

    return metadata


def load_mapping(mapping_path: Path) -> Dict[str, dict]:
    data = json.loads(mapping_path.read_text(encoding="utf-8"))
    entries = data.get("requirement_mapping", [])
    return {entry.get("old_req"): entry for entry in entries if entry.get("old_req")}


def load_ignored_requirements(ignore_csv_path: Path) -> Dict[str, str]:
    ignored: Dict[str, str] = {}
    if not ignore_csv_path.exists():
        return ignored

    with ignore_csv_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            return ignored

        old_req_key = "old_req" if "old_req" in reader.fieldnames else None
        title_key = "title" if "title" in reader.fieldnames else None
        if old_req_key is None:
            return ignored

        for row in reader:
            old_req = (row.get(old_req_key) or "").strip()
            if not OLD_REQ_RE.fullmatch(old_req):
                continue
            title = (row.get(title_key) or "").strip() if title_key else ""
            ignored[old_req] = title

    return ignored


def is_ptsb_source(source: str) -> bool:
    normalized = source.casefold()
    return "ptsb" in normalized or "gemprodt" in normalized


def is_gemspec_source(source: str) -> bool:
    return "gemspec" in source.casefold()


def is_gemf_source(source: str) -> bool:
    return "gemf" in source.casefold()


def preferred_source(sources: Set[str]) -> str:
    if not sources:
        return "Unbekannte Quelle"

    sorted_sources = sorted(sources)

    # Prioritize gemSpec sources first. gemF is fallback only when no gemSpec source exists.
    gemspec_sources = [src for src in sorted_sources if is_gemspec_source(src)]
    non_ptsb_gemspec = [src for src in gemspec_sources if not is_ptsb_source(src)]
    if non_ptsb_gemspec:
        return non_ptsb_gemspec[0]
    if gemspec_sources:
        return gemspec_sources[0]

    non_gemf_sources = [src for src in sorted_sources if not is_gemf_source(src)]
    non_ptsb_non_gemf = [src for src in non_gemf_sources if not is_ptsb_source(src)]
    if non_ptsb_non_gemf:
        return non_ptsb_non_gemf[0]
    if non_gemf_sources:
        return non_gemf_sources[0]

    non_ptsb = [src for src in sorted_sources if not is_ptsb_source(src)]
    if non_ptsb:
        return non_ptsb[0]
    return sorted_sources[0]


def aggregate_metadata_by_base(
    old_meta: Dict[str, Dict[str, Set[str]]],
) -> Dict[str, Dict[str, Set[str]]]:
    by_base: Dict[str, Dict[str, Set[str]]] = defaultdict(
        lambda: {"titles": set(), "sources": set()}
    )
    for req_id, meta in old_meta.items():
        base_id = base_requirement_id(req_id)
        by_base[base_id]["titles"].update(meta.get("titles", set()))
        by_base[base_id]["sources"].update(meta.get("sources", set()))
    return by_base


def metadata_for_req(
    req_id: str,
    old_meta: Dict[str, Dict[str, Set[str]]],
    meta_by_base: Dict[str, Dict[str, Set[str]]],
) -> Dict[str, Set[str]]:
    direct = old_meta.get(req_id)
    if direct:
        return direct

    base = base_requirement_id(req_id)
    by_base = meta_by_base.get(base)
    if by_base:
        return by_base

    return {"titles": set(), "sources": set()}


def page_url_from_mapping_file(file_path: str, base_url: str) -> str:
    # Example: igs/core/input/pagecontent/op-create-req-fd.md ->
    # {{NEW_IG_BASE_URL}}/core/op-create-req-fd.html
    path = Path(file_path)
    parts = path.parts
    ig_name = parts[1] if len(parts) > 1 else "ig"
    page_name = path.stem
    return f"{base_url}/{ig_name}/{page_name}.html"


def new_req_links(entry: dict, base_url: str) -> str:
    occs = entry.get("occurrences") or []
    if isinstance(occs, list) and occs:
        links = []
        seen = set()
        for occ in occs:
            if not isinstance(occ, dict):
                continue
            new_req = str(occ.get("new_req") or "").strip()
            file_path = str(occ.get("file") or "").strip()
            if not new_req or not file_path:
                continue
            key = (new_req, file_path)
            if key in seen:
                continue
            seen.add(key)
            url = page_url_from_mapping_file(file_path, base_url)
            links.append(f"[{new_req}]({url})")
        if links:
            return "<br/>".join(links)

    # Fallback to plain new_req values if occurrences are missing.
    raw_new = entry.get("new_req") or []
    return "<br/>".join(str(item) for item in raw_new)


def build_markdown(
    old_meta: Dict[str, Dict[str, Set[str]]],
    mapping_by_old_req: Dict[str, dict],
    ignored_requirements: Dict[str, str],
    new_req_base_url: str,
) -> str:
    meta_by_base = aggregate_metadata_by_base(old_meta)
    ignored_ids = set(ignored_requirements.keys())
    mapped_ids = sorted(req for req in mapping_by_old_req.keys() if req not in ignored_ids)

    grouped_mapped_by_source: Dict[str, List[str]] = defaultdict(list)
    for old_req in mapped_ids:
        meta = metadata_for_req(old_req, old_meta, meta_by_base)
        source = preferred_source(meta.get("sources", set()))
        grouped_mapped_by_source[source].append(old_req)

    grouped_ignored_by_source: Dict[str, List[str]] = defaultdict(list)
    for old_req in sorted(ignored_ids):
        meta = metadata_for_req(old_req, old_meta, meta_by_base)
        source = preferred_source(meta.get("sources", set()))
        grouped_ignored_by_source[source].append(old_req)

    all_sources = sorted(set(grouped_mapped_by_source.keys()) | set(grouped_ignored_by_source.keys()))
    all_sources = sorted(all_sources, key=lambda source: (is_gemf_source(source), source.casefold()))

    lines: List[str] = []
    lines.append(
        "Diese Seite hilft beim Auffinden von alten Anforderungen (PDF-Welt) in den neuen FHIR-IG Anforderungen."
    )
    lines.append("")
    lines.append(
        "Dargestellt werden je Quelldokument nur übernommene Anforderungen (gemappte Anforderungen) "
        "sowie nicht übernommene Anforderungen aus `req-ignore.csv`."
    )
    lines.append("")
    lines.append(
        f"Hinweis: Links auf neue Anforderungen verwenden aktuell den Platzhalter `{new_req_base_url}`."
    )
    lines.append("")
    for source in all_sources:
        lines.append(f"### {source}")
        lines.append("")

        lines.append("#### Übernommene Anforderungen")
        lines.append("")
        lines.append("| Alte Anforderung | Alter Titel | Neue Anforderungen | IGs |")
        lines.append("|---|---|---|---|")

        for old_req in sorted(grouped_mapped_by_source.get(source, [])):
            meta = metadata_for_req(old_req, old_meta, meta_by_base)
            entry = mapping_by_old_req.get(old_req)

            old_link = f"[{old_req}](https://gemspec.gematik.de/search/index.html?{old_req})"
            titles = "<br/>".join(sorted(meta.get("titles", set()))) or "-"

            # old_req is from mapping_by_old_req, so entry is expected.
            new_links = new_req_links(entry, new_req_base_url) or "-"
            igs = "<br/>".join(entry.get("igs") or []) or "-"

            lines.append(f"| {old_link} | {titles} | {new_links} | {igs} |")

        lines.append("")

        ignored_for_source = sorted(grouped_ignored_by_source.get(source, []))
        if ignored_for_source:
            lines.append("#### Nicht übernommene Anforderungen")
            lines.append("")
            lines.append("| Alte Anforderung | Alter Titel |")
            lines.append("|---|---|")

            for old_req in ignored_for_source:
                meta = metadata_for_req(old_req, old_meta, meta_by_base)
                old_link = f"[{old_req}](https://gemspec.gematik.de/search/index.html?{old_req})"

                title_from_ignore = ignored_requirements.get(old_req, "")
                title_values = sorted(meta.get("titles", set()))
                title = title_values[0] if title_values else title_from_ignore or "-"

                lines.append(f"| {old_link} | {title} |")

            lines.append("")

    lines.append("")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate markdown mapping page from old requirements to new IG requirements."
    )
    parser.add_argument(
        "--mapping",
        default="qa/requirement-mapping.json",
        help="Path to requirement-mapping.json (default: qa/requirement-mapping.json)",
    )
    parser.add_argument(
        "--xlsx-dir",
        default=".temp",
        help="Directory containing old requirement xlsx files (default: .temp)",
    )
    parser.add_argument(
        "--output",
        default="igs/core/input/pagecontent/requirement-mapping-old-to-new.md",
        help=(
            "Output markdown page path "
            "(default: igs/core/input/pagecontent/requirement-mapping-old-to-new.md)"
        ),
    )
    parser.add_argument(
        "--new-req-base-url",
        default="{{NEW_IG_BASE_URL}}",
        help="Placeholder base URL for links to new requirements",
    )
    parser.add_argument(
        "--req-ignore-csv",
        default="scripts/requirement-qa/config/req-ignore.csv",
        help="CSV with ignored old requirement IDs (default: scripts/requirement-qa/config/req-ignore.csv)",
    )
    args = parser.parse_args()

    mapping_path = Path(args.mapping)
    xlsx_dir = Path(args.xlsx_dir)
    output_path = Path(args.output)
    req_ignore_csv = Path(args.req_ignore_csv)

    if not mapping_path.exists():
        raise FileNotFoundError(f"Mapping file not found: {mapping_path}")
    if not xlsx_dir.is_dir():
        raise FileNotFoundError(f"XLSX directory not found: {xlsx_dir}")

    xlsx_paths = sorted(xlsx_dir.glob("*.xlsx"))
    old_meta = collect_old_req_metadata(xlsx_paths)
    mapping_by_old_req = load_mapping(mapping_path)
    ignored_requirements = load_ignored_requirements(req_ignore_csv)

    markdown = build_markdown(
        old_meta,
        mapping_by_old_req,
        ignored_requirements,
        args.new_req_base_url,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(markdown, encoding="utf-8")

    ignored_ids = set(ignored_requirements.keys())
    mapped_ids = set(mapping_by_old_req.keys()) - ignored_ids
    displayed_rows = len(mapped_ids) + len(ignored_ids)

    print(
        f"Generated mapping page with {displayed_rows} rows: {output_path}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
