#!/usr/bin/env python3
"""Verify requirement mapping coverage against A_ IDs in Excel files."""

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Dict, Iterable, List, Set

import openpyxl


OLD_REQ_ID_RE = re.compile(r"^(A_\d+)(?:-(\d+))?$")


def first_non_empty(row: tuple, indexes: List[int]) -> str:
    for idx in indexes:
        if idx >= len(row):
            continue
        value = row[idx]
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def format_source(path: Path) -> str:
    stem = path.stem
    if "_V" in stem:
        base, version = stem.rsplit("_V", 1)
        return f"{base} {version}"
    return stem


def collect_excel_ids(xlsx_paths: Iterable[Path], ptsb: bool = False) -> Dict[str, Dict[str, Set[str]]]:
    all_ids: Dict[str, Dict[str, Set[str]]] = {}
    for path in xlsx_paths:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

        sheet_names: List[str]
        if ptsb:
            if "Festlegungen" not in workbook.sheetnames:
                continue
            sheet_names = ["Festlegungen"]
        else:
            sheet_names = list(workbook.sheetnames)

        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            header = next(rows, [])
            if not header:
                continue

            index = {
                str(name).strip(): i
                for i, name in enumerate(header)
                if name is not None and str(name).strip()
            }
            if "ID" not in index:
                continue

            id_idx = index["ID"]
            title_indexes = [
                index[name]
                for name in ("Titel", "Title")
                if name in index
            ]
            content_indexes = [
                index[name]
                for name in ("Beschreibung", "Inhalt", "Content")
                if name in index
            ]

            # In PTSB mode, source should come from the sheet column.
            source_idx = index.get("Quelle (Referenz)") if ptsb else None

            for row in sheet.iter_rows(min_row=2, values_only=True):
                value = row[id_idx]
                if not value:
                    continue

                req_id = str(value).strip()
                if not req_id:
                    continue

                title = first_non_empty(row, title_indexes)
                content = first_non_empty(row, content_indexes)

                if source_idx is not None and source_idx < len(row):
                    source_value = row[source_idx]
                    source = str(source_value).strip() if source_value is not None else ""
                else:
                    source = ""
                if not source:
                    source = format_source(path)

                req_info = all_ids.setdefault(
                    req_id,
                    {"sources": set(), "titles": set(), "contents": set()},
                )
                req_info["sources"].add(source)
                if title:
                    req_info["titles"].add(title)
                if content:
                    req_info["contents"].add(content)

    return all_ids


def load_mapping(mapping_path: Path) -> List[Dict[str, List[str]]]:
    data = json.loads(mapping_path.read_text(encoding="utf-8"))
    return data.get("requirement_mapping", [])


def parse_old_req_id(req_id: str) -> tuple[str, int] | None:
    """Parse old requirement IDs and return (base_id, revision).

    Examples:
    - A_26148 -> (A_26148, 0)
    - A_26148-01 -> (A_26148, 1)
    """
    match = OLD_REQ_ID_RE.match(req_id.strip())
    if not match:
        return None

    base_id = match.group(1)
    revision = int(match.group(2)) if match.group(2) else 0
    return base_id, revision


def build_max_mapping_revision(mapping_ids: Set[str]) -> Dict[str, int]:
    max_revision_by_base: Dict[str, int] = {}
    for req_id in mapping_ids:
        parsed = parse_old_req_id(req_id)
        if not parsed:
            continue
        base_id, revision = parsed
        current = max_revision_by_base.get(base_id, -1)
        if revision > current:
            max_revision_by_base[base_id] = revision
    return max_revision_by_base


def is_excel_id_covered_by_mapping(req_id: str, max_mapping_revision: Dict[str, int]) -> bool:
    parsed = parse_old_req_id(req_id)
    if not parsed:
        return False

    base_id, excel_revision = parsed
    mapped_max_revision = max_mapping_revision.get(base_id)
    if mapped_max_revision is None:
        return False

    # Coverage rule:
    # - A_12345 is covered by A_12345 or A_12345-xx in mapping.
    # - A_12345-02 is covered only if mapping has revision >= 2 for same base.
    return mapped_max_revision >= excel_revision


def write_csv(path: Path, header: List[str], rows: List[List[str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerows(rows)


def load_single_column_csv(path: Path, preferred_columns: List[str]) -> Set[str]:
    """Load values from CSV, supporting optional headers or plain one-column files."""
    if not path.exists():
        return set()

    with path.open("r", encoding="utf-8", newline="") as handle:
        rows = [row for row in csv.reader(handle) if row]

    if not rows:
        return set()

    def clean(text: str) -> str:
        return text.strip()

    header = [clean(col).lower() for col in rows[0]]
    preferred = [c.lower() for c in preferred_columns]

    # Header mode: first row contains one of the expected column names.
    if any(col in preferred for col in header):
        value_idx = next((i for i, col in enumerate(header) if col in preferred), 0)
        data_rows = rows[1:]
    else:
        # Plain mode: first column contains values directly.
        value_idx = 0
        data_rows = rows

    values: Set[str] = set()
    for row in data_rows:
        if value_idx >= len(row):
            continue
        value = clean(row[value_idx])
        if not value or value.startswith("#"):
            continue
        values.add(value)

    return values


def source_matches_spec(source: str, allowed_specs: Set[str]) -> bool:
    for spec in allowed_specs:
        if source == spec or source.startswith(f"{spec} "):
            return True
    return False


def filter_excel_ids_by_specs(
    excel_ids: Dict[str, Dict[str, Set[str]]],
    allowed_specs: Set[str],
) -> Dict[str, Dict[str, Set[str]]]:
    if not allowed_specs:
        return excel_ids

    filtered: Dict[str, Dict[str, Set[str]]] = {}
    for req_id, info in excel_ids.items():
        if any(source_matches_spec(src, allowed_specs) for src in info.get("sources", set())):
            filtered[req_id] = info
    return filtered


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Verify requirement mapping coverage against Excel A_ IDs."
    )
    parser.add_argument(
        "--xlsx-dir",
        default=".temp",
        help="Directory containing requirement Excel files (default: .temp)",
    )
    parser.add_argument(
        "--mapping",
        default="qa/requirement-mapping.json",
        help="Path to requirement-mapping.json (default: requirement-mapping.json)",
    )
    parser.add_argument(
        "--output-dir",
        default="qa",
        help="Directory to write CSV reports (default: qa)",
    )
    parser.add_argument(
        "--description",
        action="store_true",
        help="Include requirement content/description in output CSV files",
    )
    parser.add_argument(
        "--ptsb",
        action="store_true",
        help="Only use gemProdT*.xlsx files and parse requirements from sheet 'Festlegungen'",
    )
    parser.add_argument(
        "--ignore-list",
        default="scripts/requirement-qa/config/req-ignore.csv",
        help="One-column CSV with old_req IDs to ignore for missing output",
    )
    parser.add_argument(
        "--specs-to-check",
        default="scripts/requirement-qa/config/specs-to-check.csv",
        help="One-column CSV with source specs to include in --ptsb mode",
    )
    args = parser.parse_args()

    xlsx_dir = Path(args.xlsx_dir)
    mapping_path = Path(args.mapping)
    output_dir = Path(args.output_dir)
    ignore_list_path = Path(args.ignore_list)
    specs_to_check_path = Path(args.specs_to_check)

    if not xlsx_dir.exists():
        raise FileNotFoundError(f"XLSX directory not found: {xlsx_dir}")
    if not mapping_path.exists():
        raise FileNotFoundError(f"Mapping file not found: {mapping_path}")

    output_dir.mkdir(parents=True, exist_ok=True)

    if args.ptsb:
        xlsx_paths = sorted(xlsx_dir.glob("gemProdT*.xlsx"))
    else:
        xlsx_paths = sorted(xlsx_dir.glob("*.xlsx"))

    excel_ids = collect_excel_ids(xlsx_paths, ptsb=args.ptsb)

    ignore_ids = load_single_column_csv(ignore_list_path, ["old_req", "id", "requirement", "req"])
    specs_to_check = load_single_column_csv(specs_to_check_path, ["spec", "source", "document"])

    if args.ptsb:
        excel_ids = filter_excel_ids_by_specs(excel_ids, specs_to_check)

    mapping_entries = load_mapping(mapping_path)
    mapping_ids = {entry.get("old_req") for entry in mapping_entries if entry.get("old_req")}
    max_mapping_revision = build_max_mapping_revision(mapping_ids)

    missing_ids = sorted(
        req_id
        for req_id in excel_ids.keys()
        if req_id not in ignore_ids and not is_excel_id_covered_by_mapping(req_id, max_mapping_revision)
    )

    duplicate_rows: List[List[str]] = []
    for entry in mapping_entries:
        old_req = entry.get("old_req")
        new_req = entry.get("new_req") or []
        if old_req and len(new_req) > 1:
            req_info = excel_ids.get(old_req, {})
            row = [
                old_req,
                ";".join(sorted(req_info.get("titles", set()))),
            ]
            if args.description:
                row.append("\n\n".join(sorted(req_info.get("contents", set()))))
            row.extend(
                [
                    ";".join(sorted(new_req)),
                    ";".join(sorted(entry.get("igs") or [])),
                ]
            )
            duplicate_rows.append(row)

    missing_rows = []
    for req_id in missing_ids:
        req_info = excel_ids.get(req_id, {})
        row = [
            req_id,
            ";".join(sorted(req_info.get("titles", set()))),
        ]
        if args.description:
            row.append("\n\n".join(sorted(req_info.get("contents", set()))))
        row.append(";".join(sorted(req_info.get("sources", set()))))
        missing_rows.append(row)

    missing_header = ["old_req", "title"]
    duplicate_header = ["old_req", "title"]
    if args.description:
        missing_header.append("content")
        duplicate_header.append("content")
    missing_header.append("source")
    duplicate_header.extend(["new_req", "igs"])

    write_csv(
        output_dir / "requirement-mapping-missing.csv",
        missing_header,
        missing_rows,
    )
    write_csv(
        output_dir / "requirement-mapping-duplicate.csv",
        duplicate_header,
        duplicate_rows,
    )


if __name__ == "__main__":
    main()
