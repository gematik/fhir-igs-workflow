#!/usr/bin/env python3
"""Replace A_ requirement IDs in a markdown file with requirement blocks from an Excel sheet."""

import argparse
import html
import re
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Set

import openpyxl


A_ID_PATTERN = re.compile(r"A_\d+(?:-\d+)?")
FULL_LINE_IDS_PATTERN = re.compile(r"^(A_\d+(?:-\d+)?)(\s+A_\d+(?:-\d+)?)*$")


def load_requirements(excel_path: Path, sheet_name: str) -> Dict[str, Dict[str, str]]:
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {excel_path}")

    sheet = workbook[sheet_name]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: i for i, name in enumerate(header)}

    required_columns = {"ID", "Titel", "Beschreibung", "Verbindlichkeit"}
    missing_columns = required_columns - set(index)
    if missing_columns:
        raise ValueError(f"Missing columns in Excel sheet: {', '.join(sorted(missing_columns))}")

    requirements: Dict[str, Dict[str, str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        req_id = row[index["ID"]]
        if not req_id:
            continue
        requirements[str(req_id).strip()] = {
            "title": (row[index["Titel"]] or "").strip(),
            "desc": (row[index["Beschreibung"]] or "").strip(),
            "level": (row[index["Verbindlichkeit"]] or "").strip(),
        }
    return requirements


def discover_excel_files(excel_arg: Optional[str], excel_dir: Path) -> List[Path]:
    if excel_arg:
        path = Path(excel_arg)
        return [path]

    if not excel_dir.exists() or not excel_dir.is_dir():
        raise ValueError(f"Excel directory does not exist or is not a directory: {excel_dir}")

    files = sorted(p for p in excel_dir.glob("*.xlsx") if p.is_file())
    if not files:
        raise ValueError(f"No .xlsx files found in directory: {excel_dir}")
    return files


def load_requirements_from_files(
    excel_files: Sequence[Path],
    sheet_name: str,
    required_ids: Optional[Set[str]] = None,
) -> Dict[str, Dict[str, str]]:
    merged: Dict[str, Dict[str, str]] = {}
    accepted_files: List[Path] = []

    for excel_file in excel_files:
        try:
            requirements = load_requirements(excel_file, sheet_name)
        except Exception as exc:
            print(f"Skipping non-conforming Excel file {excel_file}: {exc}")
            continue

        accepted_files.append(excel_file)
        for req_id, req in requirements.items():
            if required_ids is not None and req_id not in required_ids:
                continue

            if req_id in merged and merged[req_id] != req:
                print(
                    f"Warning: conflicting definition for {req_id} in {excel_file}; "
                    "keeping the first occurrence"
                )
                continue
            merged.setdefault(req_id, req)

    if not accepted_files:
        files_list = ", ".join(str(p) for p in excel_files)
        raise ValueError(
            "No conforming Excel files found. "
            f"Checked files: {files_list}. Expected sheet '{sheet_name}' with columns "
            "ID, Titel, Beschreibung, Verbindlichkeit."
        )

    print(f"Using {len(accepted_files)} conforming Excel file(s).")
    return merged


def collect_required_ids(markdown_content: str) -> Set[str]:
    required_ids: Set[str] = set()
    for line in markdown_content.splitlines():
        stripped = line.strip()
        if stripped and FULL_LINE_IDS_PATTERN.fullmatch(stripped):
            required_ids.update(A_ID_PATTERN.findall(stripped))
    return required_ids


def map_conformance(level: str) -> str:
    level = level.upper()
    if level == "MUSS":
        return "SHALL"
    if level == "SOLL":
        return "SHOULD"
    if level in {"KANN", "DARF"}:
        return "MAY"
    return "SHALL"


def format_description(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = " ".join(text.split())
    return html.escape(text, quote=False)


def build_block(req_id: str, req: Dict[str, str], actor: str, test_procedure: str) -> str:
    title = req["title"].replace('"', "&quot;")
    conformance = map_conformance(req["level"])
    description = format_description(req["desc"])
    return (
        f"<!-- {req_id} -->\n"
        f"<requirement conformance=\"{conformance}\" title=\"{title}\">\n"
        "    <meta lockversion=\"false\"/>\n"
        f"    <actor name=\"{actor}\">\n"
        f"        <testProcedure id=\"{test_procedure}\"/>\n"
        "    </actor>\n"
        f"     {description}\n"
        "</requirement>\n"
    )


def replace_in_markdown(content: str, requirements: Dict[str, Dict[str, str]], actor: str, test_procedure: str) -> str:
    lines = content.splitlines()
    output_lines: List[str] = []

    for line in lines:
        stripped = line.strip()
        if not stripped or not FULL_LINE_IDS_PATTERN.fullmatch(stripped):
            output_lines.append(line)
            continue

        ids = A_ID_PATTERN.findall(stripped)

        blocks = []
        for req_id in ids:
            if req_id not in requirements:
                raise ValueError(f"Requirement ID '{req_id}' not found in Excel")
            blocks.append(build_block(req_id, requirements[req_id], actor, test_procedure).rstrip())

        output_lines.append("\n\n".join(blocks))

    return "\n".join(output_lines).rstrip() + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Replace A_ requirement IDs in a markdown file.")
    parser.add_argument("--markdown", required=True, help="Path to the markdown file to update")
    parser.add_argument("--excel", help="Path to a single Excel file with requirements")
    parser.add_argument(
        "--excel-dir",
        default=".temp",
        help="Directory to scan for .xlsx files when --excel is not provided",
    )
    parser.add_argument("--sheet", default="Festlegungen", help="Sheet name in the Excel file")
    parser.add_argument("--actor", default="E-Rezept-Fachdienst", help="Actor name for the requirement block")
    parser.add_argument("--test-procedure", default="Produkttest", help="Test procedure id")

    args = parser.parse_args()

    markdown_path = Path(args.markdown)
    content = markdown_path.read_text(encoding="utf-8")
    required_ids = collect_required_ids(content)

    excel_files = discover_excel_files(args.excel, Path(args.excel_dir))
    requirements = load_requirements_from_files(excel_files, args.sheet, required_ids)

    updated = replace_in_markdown(content, requirements, args.actor, args.test_procedure)
    markdown_path.write_text(updated, encoding="utf-8")


if __name__ == "__main__":
    main()
