#!/usr/bin/env python3
"""Export latest requirement releases from all IGs into one Excel workbook.

For each IG under ./igs, this script selects the highest semver release directory
from .igtools/releases, reads requirement YAML files, and creates one worksheet
in a single XLSX file.

Columns per worksheet:
- id
- title
- conformance
- test_procedure

The test_procedure column uses actor/test-procedure display labels. When a
requirement has multiple actors, each actor gets its own bullet point.
"""

from __future__ import annotations

import argparse
import csv
import re
import subprocess
import sys
from pathlib import Path
from typing import Any

import openpyxl
import yaml
from openpyxl.styles import Alignment, Font


SEMVER_DIR_RE = re.compile(
    r"^(?P<major>\d+)[._](?P<minor>\d+)[._](?P<patch>\d+)(?P<suffix>.*)$"
)
REQUIREMENT_FILE_RE = re.compile(r"^IG-.*\.ya?ml$", re.IGNORECASE)

# Aligned with scripts/requirement-qa/check_requirement_quality.py semantics.
CONFORMANCE_TRANSLATIONS: dict[str, str] = {
    "SHALL": "MUSS",
    "MUST": "MUSS",
    "SHALL NOT": "DARF NICHT",
    "MUST NOT": "DARF NICHT",
    "SHOULD": "SOLL",
    "SHOULD NOT": "SOLL NICHT",
    "MAY": "KANN",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Export latest release requirements from all IGs into one XLSX workbook."
        )
    )
    parser.add_argument(
        "--igs-root",
        default="igs",
        help="Directory that contains IG folders (default: igs)",
    )
    parser.add_argument(
        "--output",
        default="requirements-latest-releases.xlsx",
        help="Output XLSX path (default: requirements-latest-releases.xlsx)",
    )
    parser.add_argument(
        "--skip-polarion-mapping",
        action="store_true",
        help="Do not run `igtools polarion-mapping` before export",
    )
    return parser.parse_args()


def run_polarion_mapping(repo_root: Path) -> None:
    try:
        subprocess.run(
            ["igtools", "polarion-mapping"],
            cwd=repo_root,
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    except (FileNotFoundError, subprocess.CalledProcessError):
        print(
            "Warning: Could not run `igtools polarion-mapping`. "
            "Continuing with repository mapping files.",
            file=sys.stderr,
        )


def load_mapping_csv(path: Path) -> dict[str, str]:
    mapping: dict[str, str] = {}
    if not path.exists():
        return mapping

    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        for row in reader:
            if len(row) < 3:
                continue
            key = row[1].strip()
            label = row[2].strip()
            if key:
                mapping[key] = label or key
    return mapping


def parse_semver_dir(name: str) -> tuple[int, int, int, str] | None:
    match = SEMVER_DIR_RE.match(name)
    if not match:
        return None
    major = int(match.group("major"))
    minor = int(match.group("minor"))
    patch = int(match.group("patch"))
    suffix = match.group("suffix") or ""
    return major, minor, patch, suffix


def semver_display_from_dir(name: str) -> str:
    parsed = parse_semver_dir(name)
    if parsed is None:
        return name
    major, minor, patch, suffix = parsed
    base = f"{major}.{minor}.{patch}"
    if suffix:
        return f"{base}{suffix}"
    return base


def latest_release_dir(releases_dir: Path) -> Path | None:
    candidates: list[tuple[tuple[int, int, int, int, str], Path]] = []
    for child in releases_dir.iterdir():
        if not child.is_dir():
            continue
        parsed = parse_semver_dir(child.name)
        if parsed is None:
            continue
        major, minor, patch, suffix = parsed
        # Plain versions sort after pre-release variants with a suffix.
        is_plain = 1 if suffix == "" else 0
        candidates.append(((major, minor, patch, is_plain, suffix), child))

    if not candidates:
        return None

    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def sanitize_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name)
    cleaned = cleaned[:31] or "Sheet"
    candidate = cleaned
    i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        i += 1
    used.add(candidate)
    return candidate


def maybe_map_test_procedure(value: str, procedure_map: dict[str, str]) -> str:
    text = value.strip()
    if text in procedure_map:
        return procedure_map[text]
    return text


def translate_conformance(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = re.sub(r"\s+", " ", text.upper())
    return CONFORMANCE_TRANSLATIONS.get(normalized, text)


def to_actor_test_procedure_text(
    requirement: dict[str, Any],
    actor_map: dict[str, str],
    procedure_map: dict[str, str],
) -> str:
    actors = requirement.get("actor") or []
    if not isinstance(actors, list):
        actors = []

    test_procedures = requirement.get("test_procedures") or {}
    if not isinstance(test_procedures, dict):
        test_procedures = {}

    lines: list[str] = []
    multiple_actors = len(actors) > 1

    for actor_key in actors:
        if not isinstance(actor_key, str):
            continue
        actor_label = actor_map.get(actor_key, actor_key)
        procedure_values = test_procedures.get(actor_key, [])

        if isinstance(procedure_values, str):
            procedure_values = [procedure_values]
        if not isinstance(procedure_values, list):
            procedure_values = []

        mapped = [
            maybe_map_test_procedure(str(item), procedure_map)
            for item in procedure_values
            if str(item).strip()
        ]
        procedures_text = ", ".join(mapped) if mapped else "-"

        if multiple_actors:
            lines.append(f"• {actor_label}: {procedures_text}")
        else:
            lines.append(f"{actor_label}: {procedures_text}")

    if not lines:
        return ""
    return "\n".join(lines)


def read_requirement_yaml(path: Path) -> dict[str, Any] | None:
    try:
        data = yaml.safe_load(path.read_text(encoding="utf-8"))
    except (OSError, yaml.YAMLError):
        return None
    if isinstance(data, dict):
        return data
    return None


def collect_release_requirements(release_dir: Path) -> list[dict[str, Any]]:
    requirements: list[dict[str, Any]] = []
    for file in sorted(release_dir.iterdir()):
        if not file.is_file() or not REQUIREMENT_FILE_RE.match(file.name):
            continue
        req = read_requirement_yaml(file)
        if req is None:
            continue
        if "key" not in req:
            continue
        requirements.append(req)
    requirements.sort(key=lambda req: str(req.get("key", "")))
    return requirements


def configure_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    ws.append(["id", "title", "conformance", "test_procedure", "release_status"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")
    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 85
    ws.column_dimensions["E"].width = 18


def append_requirement_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    requirement: dict[str, Any],
    actor_map: dict[str, str],
    procedure_map: dict[str, str],
) -> None:
    test_procedure_text = to_actor_test_procedure_text(
        requirement=requirement,
        actor_map=actor_map,
        procedure_map=procedure_map,
    )
    ws.append(
        [
            requirement.get("key", ""),
            requirement.get("title", ""),
            translate_conformance(requirement.get("conformance", "")),
            test_procedure_text,
            requirement.get("release_status", ""),
        ]
    )
    row_idx = ws.max_row
    ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True, vertical="top")


def export_latest_requirements(
    repo_root: Path,
    igs_root: Path,
    output_path: Path,
    run_mapping: bool,
) -> int:
    if run_mapping:
        run_polarion_mapping(repo_root)

    actor_map = load_mapping_csv(repo_root / "spec_tools" / "polarion_actors.csv")
    procedure_map = load_mapping_csv(
        repo_root / "spec_tools" / "polarion_test_procedures.csv"
    )

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    used_sheet_names: set[str] = set()
    ig_count = 0

    for ig_dir in sorted(igs_root.iterdir()):
        if not ig_dir.is_dir():
            continue

        releases_dir = ig_dir / ".igtools" / "releases"
        if not releases_dir.is_dir():
            continue

        latest_dir = latest_release_dir(releases_dir)
        if latest_dir is None:
            continue

        requirements = collect_release_requirements(latest_dir)
        if not requirements:
            continue

        release_display = semver_display_from_dir(latest_dir.name)
        sheet_name = sanitize_sheet_name(
            f"{ig_dir.name} {release_display}",
            used_sheet_names,
        )
        ws = workbook.create_sheet(sheet_name)
        configure_sheet(ws)

        for requirement in requirements:
            append_requirement_row(ws, requirement, actor_map, procedure_map)

        ig_count += 1
        print(
            f"Added sheet '{sheet_name}' from {latest_dir.relative_to(repo_root)} "
            f"with {len(requirements)} requirements."
        )

    if ig_count == 0:
        print("No IG releases with requirement YAML files found.", file=sys.stderr)
        return 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Wrote workbook: {output_path}")
    return 0


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parent.parent
    igs_root = (repo_root / args.igs_root).resolve()
    output_path = (repo_root / args.output).resolve()

    if not igs_root.is_dir():
        print(f"IG root does not exist: {igs_root}", file=sys.stderr)
        return 1

    return export_latest_requirements(
        repo_root=repo_root,
        igs_root=igs_root,
        output_path=output_path,
        run_mapping=not args.skip_polarion_mapping,
    )


if __name__ == "__main__":
    raise SystemExit(main())
