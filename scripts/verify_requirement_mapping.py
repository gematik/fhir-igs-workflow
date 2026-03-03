#!/usr/bin/env python3
"""Verify requirement mapping coverage against A_ IDs in Excel files."""

import argparse
import csv
import json
from pathlib import Path
from typing import Dict, Iterable, List, Set

import openpyxl


def first_non_empty(row: tuple, indexes: List[int]) -> str:
    for idx in indexes:
        if idx >= len(row):
            continue
        value = row[idx]
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def format_source(path: Path) -> str:
    stem = path.stem
    if "_V" in stem:
        base, version = stem.rsplit("_V", 1)
        return f"{base} {version}"
    return stem


def collect_excel_ids(xlsx_paths: Iterable[Path]) -> Dict[str, Dict[str, Set[str]]]:
    all_ids: Dict[str, Dict[str, Set[str]]] = {}
    for path in xlsx_paths:
        source = format_source(path)
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            header = next(rows, [])
            if not header:
                continue
            index = {name: i for i, name in enumerate(header) if name is not None}
            if "ID" not in index:
                continue
            id_idx = index["ID"]
            title_indexes = [
                index[name]
                for name in ("Titel", "Title")
                if name in index
            ]
            content_indexes = [
                index[name]
                for name in ("Beschreibung", "Inhalt", "Content")
                if name in index
            ]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                value = row[id_idx]
                if not value:
                    continue
                req_id = str(value).strip()
                if req_id:
                    title = first_non_empty(row, title_indexes)
                    content = first_non_empty(row, content_indexes)
                    req_info = all_ids.setdefault(
                        req_id,
                        {"sources": set(), "titles": set(), "contents": set()},
                    )
                    req_info["sources"].add(source)
                    if title:
                        req_info["titles"].add(title)
                    if content:
                        req_info["contents"].add(content)
    return all_ids


def load_mapping(mapping_path: Path) -> List[Dict[str, List[str]]]:
    data = json.loads(mapping_path.read_text(encoding="utf-8"))
    return data.get("requirement_mapping", [])


def write_csv(path: Path, header: List[str], rows: List[List[str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Verify requirement mapping coverage against Excel A_ IDs."
    )
    parser.add_argument(
        "--xlsx-dir",
        default=".temp",
        help="Directory containing requirement Excel files (default: .temp)",
    )
    parser.add_argument(
        "--mapping",
        default="qa/requirement-mapping.json",
        help="Path to requirement-mapping.json (default: requirement-mapping.json)",
    )
    parser.add_argument(
        "--output-dir",
        default="qa",
        help="Directory to write CSV reports (default: qa)",
    )
    parser.add_argument(
        "--description",
        action="store_true",
        help="Include requirement content/description in output CSV files",
    )
    args = parser.parse_args()

    xlsx_dir = Path(args.xlsx_dir)
    mapping_path = Path(args.mapping)
    output_dir = Path(args.output_dir)

    if not xlsx_dir.exists():
        raise FileNotFoundError(f"XLSX directory not found: {xlsx_dir}")
    if not mapping_path.exists():
        raise FileNotFoundError(f"Mapping file not found: {mapping_path}")

    output_dir.mkdir(parents=True, exist_ok=True)

    xlsx_paths = sorted(xlsx_dir.glob("*.xlsx"))
    excel_ids = collect_excel_ids(xlsx_paths)

    mapping_entries = load_mapping(mapping_path)
    mapping_ids = {entry.get("old_req") for entry in mapping_entries if entry.get("old_req")}

    missing_ids = sorted(set(excel_ids.keys()) - mapping_ids)

    duplicate_rows: List[List[str]] = []
    for entry in mapping_entries:
        old_req = entry.get("old_req")
        new_req = entry.get("new_req") or []
        if old_req and len(new_req) > 1:
            req_info = excel_ids.get(old_req, {})
            row = [
                old_req,
                ";".join(sorted(req_info.get("titles", set()))),
            ]
            if args.description:
                row.append("\n\n".join(sorted(req_info.get("contents", set()))))
            row.extend(
                [
                    ";".join(sorted(new_req)),
                    ";".join(sorted(entry.get("igs") or [])),
                ]
            )
            duplicate_rows.append(row)

    missing_rows = []
    for req_id in missing_ids:
        req_info = excel_ids.get(req_id, {})
        row = [
            req_id,
            ";".join(sorted(req_info.get("titles", set()))),
        ]
        if args.description:
            row.append("\n\n".join(sorted(req_info.get("contents", set()))))
        row.append(";".join(sorted(req_info.get("sources", set()))))
        missing_rows.append(row)

    missing_header = ["old_req", "title"]
    duplicate_header = ["old_req", "title"]
    if args.description:
        missing_header.append("content")
        duplicate_header.append("content")
    missing_header.append("source")
    duplicate_header.extend(["new_req", "igs"])

    write_csv(
        output_dir / "requirement-mapping-missing.csv",
        missing_header,
        missing_rows,
    )
    write_csv(
        output_dir / "requirement-mapping-duplicate.csv",
        duplicate_header,
        duplicate_rows,
    )


if __name__ == "__main__":
    main()
